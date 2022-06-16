from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import cell
from openpyxl.styles import Alignment #, named_styles
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from selenium.webdriver.common.by import By
import sys

# Provide branch url. Example:

# Aktau 'https://2gis.kz/aktau/branches/70000001028456934'
# Aktobe 'https://2gis.kz/aktobe/branches/70000001032326074'
# Almaty: 'https://2gis.kz/almaty/branches/9429948590733484'
# Astana 'https://2gis.kz/nur_sultan/branches/70000001018099222'
# Atyrau 'https://2gis.kz/atyrau/branches/70000001035053200'
# Karaganda 'https://2gis.kz/karaganda/branches/11822485892766619'
# Kokshetau 'https://2gis.kz/kokshetau/branches/70000001033428236'
# Kostanai 'https://2gis.kz/kostanay/branches/70000001035235895'
# Pavlodar 'https://2gis.kz/pavlodar/branches/70000001018654399'
# Petropavlovsk 'https://2gis.kz/petropavlovsk/branches/70000001038197192'
# Semey No data
# Taraz 'https://2gis.kz/taraz/branches/70000001051349477'
# Uralsk 'https://2gis.kz/uralsk/branches/70000001025080018'
# Ust Kamenogorsk 'https://2gis.kz/ustkam/branches/12807648311251661'
# Shymkent 'https://2gis.kz/shymkent/branches/70000001025934880'




def print_instructions():
    print('''
    Программа предназначена для поиска ссылок на филиалы организации (точку продаж).
    Вставьте прямую ссылку на страницу с филиалами организации без пробелов и других знаков. Пример: 'https://2gis.kz/almaty/branches/9429948590733484'
    Если Ctrl+V в терминале не работает, попробуйте нажать правую кнопку мыши в месте для вставки после того, как скопировали ссылку.

    У Вас должен быть устанолен браузер Firefox, потому что программа взаимодействует с ним для поиска дынныъ. 
    Чтобы управлять Firefox с помощью программного кода используется 'geckodriver' https://github.com/mozilla/geckodriver/releases.
    Если браузер не работает возможно папку с geckodriver.exe нужно добавить в PATH.
    Версия geckodriver.exe должна соответсвовать версии Firefox на компьютере.

    Файл будет сохранен в ту же папку, где была запущена программа.
    ''')

def get_links():
    user_input = input('Введите ссылки через запятую, нажмите Enter: ')
    direct_links_to_branches = user_input.split(",")
    return direct_links_to_branches

def go_to_link(browser, direct_link_to_branches):
    browser.get(direct_link_to_branches)
    return browser

def get_2gis_brach_count(browser):
    branch_count = browser.find_elements(By.CLASS_NAME, '_1p8iqzw')[1].text # Target text example "10 филиалов организации"
    return branch_count

def get_links_to_branches(browser):
    # RETURNS NOT ONLY LINKS BUT SOME IRRELEVANT DATA
    # LOOPS UNTIL EACH ELEMENT CONTAINING COMPANY LINK IS FOUND
    while True:
        try:
            
            content_blocks = browser.find_elements_by_class_name("_1rehek")     #  search for something like: <a href="/pavlodar/firm/70000001030269073" class="_pbcct4"><span class="_hc69qa"><span class="_oqoid">проспект Нурсултана Назарбаева, 10</span></span></a>
            browser.execute_script("arguments[0].scrollIntoView();"  , content_blocks[-2] ) # may not work with "-1"
            #code returns error if  button not found on page
            add_company_button = browser.find_element_by_class_name('_6gpij7j')
            
        except NoSuchElementException as e:
            print(e)
            continue
        except Exception as e:
            print(e)
            break
        else:
            # scroll to add_company_button  
            browser.execute_script("arguments[0].scrollIntoView();"  , add_company_button )
            content_blocks = browser.find_elements_by_class_name("_1rehek") 
            break
    return content_blocks 

def get_clean_links(content_blocks):  
    raw_hrefs = []
    clean_hrefs = []
    
    for block in content_blocks:
        href = block.get_attribute('href')
        raw_hrefs.append(href)

    for href in raw_hrefs:
        if 'firm' in str(href):
            clean_hrefs.append(href)

    for c in clean_hrefs:
        print(c)         

    return clean_hrefs   

def write_data(all_hrefs):
    wb_urls = Workbook()
    ws_urls = wb_urls.active

    row = 1    
    for href_list in all_hrefs:
        for href in href_list:
            ws_urls.cell(row = row, column = 1, value=f'{href}')
            row += 1
    return wb_urls

# SAVE FILE
def save(browser, wb_urls):
    cwd = os.getcwd()
    splitedcwd = os.path.split(cwd)
    start = splitedcwd[0]
    end = splitedcwd[1]
    date = datetime.today().strftime('%Y-%m-%d %H-%M-%S')
    # city_name = direct_link_to_branches.split('/')
    filename =  f'2gis_urls' # {city_name[3]}
    filetype = '.xlsx'
    filepath = os.path.join(start + '/' + end + '/' + filename + ' ' + date + filetype)
    wb_urls.save(filepath)
    wb_urls.close()
    browser.quit() 

    
    print(f'Данные сохранены в папку "{start}\\{end}\\", файл "{filename} {date}{filetype}"', end='\n\n')
    input('Нажмите Enter для выхода.')

    sys.exit()

if __name__ == '__main__':
    all_links = []
    print_instructions()
    direct_links_to_branches = get_links()
    browser = webdriver.Firefox()
    for direct_link in direct_links_to_branches:
        browser = go_to_link(browser, direct_link)
        branch_count = get_2gis_brach_count(browser)
        content_blocks = get_links_to_branches(browser)
        clean_hrefs = get_clean_links(content_blocks)
        print('\nНайдено',len(clean_hrefs), 'ссылок из' , branch_count)
        all_links.append(clean_hrefs)

    wb_urls = write_data(all_links)
    save(browser, wb_urls)